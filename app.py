import os
import json
import logging
import tempfile
import re
from pathlib import Path
from flask import Flask, request, jsonify, render_template
import openpyxl

# LangChain – document loading & splitting
from langchain_community.document_loaders import Docx2txtLoader, TextLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter

# Load .env file if python-dotenv is installed (optional)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# AI client – provider-agnostic, configured via environment variables
import ai_client

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)-8s] %(name)s — %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
PROMPTS_DIR = BASE_DIR / "prompts"

# ── Chunking config & Constants ──────────────────────────────────────────────
BRD_CHUNK_SIZE   = 4_000
EPIC_CHUNK_SIZE  = 2_000
CHUNK_OVERLAP    = 100
MD_TO_SP_RATIO   = 1.3

# ── Startup banner ─────────────────────────────────────────────────────────
_cfg = ai_client.get_config()
logger.info("=" * 60)
logger.info("  Project Dashboard — starting up")
logger.info("  AI endpoint : %s", _cfg["base_url"])
logger.info("  Model       : %s", _cfg["model"])
logger.info("  Timeout     : %ss    Max tokens: %s",
            _cfg["timeout"], _cfg["max_tokens"] or "provider default")
logger.info("=" * 60)


# ── Helpers ────────────────────────────────────────────────────────────────

def load_prompt(filename: str) -> str:
    """Read a prompt file from the prompts/ directory."""
    return (PROMPTS_DIR / filename).read_text(encoding="utf-8").strip()


def extract_excel_epics(path: str) -> str:
    """
    Parse an Excel file (.xlsx/.xls) and extract Epic items.

    Searches every sheet for a header row that contains both a 'Summary'
    and a 'Description' column (case-insensitive, any position).
    Returns a plain-text block with one entry per epic row:

        Epic 1 — <Summary>
        <Description>

    Rows where both Summary and Description are blank are skipped.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    entries: list[str] = []

    for sheet in wb.worksheets:
        # Scan rows to find the header row
        header_row_idx: int | None = None
        summary_col: int | None = None
        desc_col: int | None = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Look for the header row by searching for both target columns
            row_lower = [
                str(cell).strip().lower() if cell is not None else ""
                for cell in row
            ]
            has_summary = "summary" in row_lower
            has_desc    = "description" in row_lower
            if has_summary and has_desc:
                summary_col     = row_lower.index("summary")
                desc_col        = row_lower.index("description")
                header_row_idx  = row_idx
                break   # found header — stop scanning

        if header_row_idx is None:
            continue   # this sheet has no matching header; try next

        # Extract data rows (everything after the header)
        epic_num = 0
        for row in sheet.iter_rows(
            min_row=header_row_idx + 1, values_only=True
        ):
            summary = str(row[summary_col]).strip() if row[summary_col] is not None else ""
            desc    = str(row[desc_col]).strip()    if row[desc_col]    is not None else ""
            if not summary and not desc:
                continue
            epic_num += 1
            parts = [f"Epic {epic_num} — {summary}" if summary else f"Epic {epic_num}"]
            if desc:
                parts.append(desc)
            entries.append("\n".join(parts))

        if entries:
            break   # found data in this sheet; don't scan further sheets

    wb.close()

    if not entries:
        raise ValueError(
            "No 'Summary' and 'Description' columns found in the Excel file. "
            "Please ensure the file has a header row with those column names."
        )

    return "\n\n".join(entries)


def extract_excel_tabular(path: str, max_cols: int = 3) -> str:
    """
    Parse an Excel file (.xlsx/.xls) and extract a generic tabular representation.
    Iterates through all sheets and extracts up to `max_cols` columns per row.
    Returns a plain-text block for AI metric extraction.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    entries: list[str] = []
    
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cols = [str(c).strip() if c is not None else "" for c in row[:max_cols]]
            # If at least one column has data
            if any(cols):
                entries.append(" | ".join(cols))
                
    wb.close()
    
    if not entries:
        raise ValueError("No data found in the Excel file.")
        
    return "\n".join(entries)


def extract_text(tmp_path: str, ext: str, mime: str, file_type: str = "default") -> str:
    """Load a document with the appropriate loader and return plain text."""
    if ext in (".xlsx", ".xls"):
        if file_type in ("jira", "bitbucket"):
            return extract_excel_tabular(tmp_path)
        return extract_excel_epics(tmp_path)
    elif ext in (".docx", ".doc"):
        loader = Docx2txtLoader(tmp_path)
    else:
        loader = TextLoader(tmp_path, encoding="utf-8", autodetect_encoding=True)

    docs = loader.load()
    return "\n\n".join(d.page_content for d in docs)


def save_and_extract(file_storage, file_type: str = "default") -> str:
    """Save a Flask FileStorage object to a temp file, extract its text, delete it."""
    if not file_storage or not file_storage.filename:
        return ""
    ext = os.path.splitext(file_storage.filename)[1].lower()
    mime = file_storage.mimetype or "application/octet-stream"
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            file_storage.save(tmp.name)
            tmp_path = tmp.name
        return extract_text(tmp_path, ext, mime, file_type)
    except Exception as exc:
        logger.warning("Could not extract text from %s: %s", file_storage.filename, exc)
        return ""
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def chunk_text(text: str, chunk_size: int = EPIC_CHUNK_SIZE) -> list:
    """Split text into overlapping chunks using LangChain's RecursiveCharacterTextSplitter."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    return splitter.split_text(text)


def ai_respond(system_prompt: str, conversation: list, user_message: str) -> str:
    """
    Delegate to ai_client.chat().
    conversation is a list of {"role": "user"|"assistant", "content": str}.
    """
    return ai_client.chat(system_prompt, conversation, user_message)


def _extract_json_block(text: str, is_array: bool = False) -> dict | list | None:
    """Finds the last valid JSON block in the text. Matches markdown fences first, then braces."""
    # Try to find markdown blocks first
    matches = re.findall(r'```(?:json)?(.*?)```', text, re.DOTALL | re.IGNORECASE)
    for match in reversed(matches):
        try:
            parsed = json.loads(match.strip())
            if isinstance(parsed, list if is_array else dict):
                return parsed
        except json.JSONDecodeError:
            pass

    # Fallback to brace matching
    start_char = '[' if is_array else '{'
    end_char = ']' if is_array else '}'
    
    blocks = []
    level = 0
    start = -1
    in_string = False
    escape = False
    for i, c in enumerate(text):
        if c == '"' and not escape:
            in_string = not in_string
        elif c == '\\' and not escape:
            escape = True
            continue
        else:
            escape = False
            
        if not in_string:
            if c == start_char:
                if level == 0:
                    start = i
                level += 1
            elif c == end_char:
                if level > 0:
                    level -= 1
                    if level == 0 and start != -1:
                        blocks.append(text[start:i+1])
                        start = -1

    for block in reversed(blocks):
        try:
            parsed = json.loads(block)
            if isinstance(parsed, list if is_array else dict):
                return parsed
        except json.JSONDecodeError:
            continue
            
    return None

def _parse_json_response(text: str) -> dict:
    """Strip markdown fences and parse JSON from a model response."""
    raw = text.strip()
    if not raw:
        raise ValueError("AI returned an empty response instead of JSON.")
    
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass
        
    extracted = _extract_json_block(raw, is_array=False)
    if extracted is not None:
        return extracted
            
    raise ValueError(f"AI returned invalid JSON. Response snippet: {raw[:200]}")


# System prompt used only for the BRD summarisation pass (terse, no JSON output)
_BRD_READER_SYSTEM = (
    "You are reading a Business Requirements Document section by section. "
    "For each section, extract the 2-3 most important technical scope items "
    "relevant for in-house custom software development effort. "
    "CRITICAL: Explicitly ignore any mention of purchased, off-the-shelf, or third-party applications "
    "that do not require custom development — we only estimate custom development effort. "
    "Be extremely brief — one short line per item. No preamble, no numbering."
)


def run_estimation(
    system_prompt: str,
    brd_chunks: list,        # may be empty when BRD not uploaded
    epic_chunks: list,
    finalize_basic_msg: str,
    finalize_contextual_msg: str,
) -> tuple[dict, dict]:
    """
    Three-phase estimation:

    Phase 0 (BRD summarisation — only if brd_chunks provided):
      Each BRD chunk is sent stateless to _BRD_READER_SYSTEM.
      The model extracts 2-3 key scope bullets per chunk.
      All bullets are concatenated into a compact context block
      that is appended to system_prompt before Epic List processing.

    Phase 1 (Epic List finalise — basic estimate):
      The Epic chunks are concatenated and sent in one single prompt alongside FINALIZE_BASIC 
      to ensure the model evaluates the FULL project scope accurately.

    Phase 3 (Contextual follow-up):
      Reuses history from Phase 2 only; applies org rules.

    Returns (basic_result, contextual_result).
    """
    # Determine total number of AI calls for logging
    num_brd_calls = len(brd_chunks) if brd_chunks else 0
    num_epic_calls = len(epic_chunks) if epic_chunks else 0
    # +1 for final basic estimate, +1 for contextual follow-up
    total_calls = num_brd_calls + num_epic_calls + 2
    call_idx = 1

    # ── Phase 0: BRD map-reduce summarisation ─────────────────────────────
    enriched_system = system_prompt
    if brd_chunks:
        total_brd = len(brd_chunks)
        logger.info("BRD summarisation: %d chunks", total_brd)
        brd_bullets: list[str] = []
        for i, chunk in enumerate(brd_chunks):
            logger.info("API call %d/%d – BRD Section %d/%d", call_idx, total_calls, i + 1, total_brd)
            msg = (
                f"[BRD Section {i + 1} of {total_brd}]\n\n"
                f"{chunk}\n\n"
                "Summarise the key scope/technical items from this section "
                "(2-3 lines max)."
            )
            summary = ai_respond(_BRD_READER_SYSTEM, [], msg).strip()
            if summary:
                brd_bullets.append(summary)
            logger.info("BRD section %d/%d summarised (%d chars)",
                        i + 1, total_brd, len(summary))
            call_idx += 1

        if brd_bullets:
            brd_context = "\n".join(brd_bullets)
            enriched_system = (
                system_prompt
                + "\n\n────────────────────────────────────────────────\n"
                f"BRD CONTEXT ({total_brd} sections summarised):\n"
                f"{brd_context}\n"
                "────────────────────────────────────────────────"
            )
            logger.info("BRD context injected: %d chars total", len(brd_context))

    # ── Phase 1 & 2: Epic List (skipped if none uploaded) ─────────────────
    if epic_chunks:
        # Combine all Epic list chunks so the model actually reads the FULL scope
        # at once instead of discarding them statelessly.
        total_epic = len(epic_chunks)
        epic_bullets = "\n\n".join([f"--- EPIC LIST PART {i+1} ---\n{chunk}" for i, chunk in enumerate(epic_chunks)])
        
        final_msg = (
            f"Here is the complete Epic List (broken into {total_epic} parts for readability):\n\n"
            f"{epic_bullets}\n\n"
            f"---\n{finalize_basic_msg}"
        )
        
        # Advance call index to simulate the old loop for accurate logging
        logger.info("Combined %d epic chunks into a single final prompt", total_epic)
        call_idx += total_epic - 1
    else:
        # BRD-only mode: ask the model to estimate directly from the BRD context
        logger.info("No Epic List — estimating from BRD context only")
        final_msg = (
            "You have read the Business Requirements Document (summarised above).\n"
            f"---\n{finalize_basic_msg}"
        )

    logger.info("API call %d/%d – Finalize Basic", call_idx, total_calls)
    basic_text = ai_respond(enriched_system, [], final_msg)
    call_idx += 1
    
    # If the model stubbornly just replies OK instead of giving the JSON, ask it again statelessly.
    if basic_text and basic_text.strip().upper() in ["OK", "OK.", '"OK"', "'OK'"]:
        logger.warning("Model replied with OK instead of JSON. Retrying statelessly.")
        retry_msg = final_msg + "\n\nCRITICAL: You just replied 'OK' and ignored the instruction. You MUST output ONLY the valid JSON structure requested, containing your estimate for all Epics."
        basic_text = ai_respond(enriched_system, [], retry_msg)
        logger.info("Retry response length: %d chars", len(basic_text))

    logger.info("Finalize basic response: %d chars", len(basic_text))

    if not basic_text:
        raise ValueError("AI returned an empty response for the final basic estimate.")

    basic_result = _parse_json_response(basic_text)
    logger.info("Basic result: %s", basic_result)
    # ── Phase 3: contextual follow-up ──────────────────────────────────
    logger.info("API call %d/%d – Contextual Follow‑up", call_idx, total_calls)
    
    # We MUST pass the basic_text we just generated back to the model as 'history'
    # so the model knows what it previously estimated before applying org rules.
    history = [
        {"role": "user",      "content": final_msg},
        {"role": "assistant", "content": basic_text},
    ]
    contextual_text = ai_respond(enriched_system, history, finalize_contextual_msg)
    logger.info("Contextual response text: %d chars", len(contextual_text))
    contextual_result = _parse_json_response(contextual_text)
    logger.info("Contextual result: %s", contextual_result)

    return basic_result, contextual_result



def generate_insights(available_data: dict) -> list:
    """
    Generate 3-5 key insights via a single LM Studio call.
    """
    lines = []
    for label, md in available_data.items():
        try:
            md_val = float(md)
        except (ValueError, TypeError):
            continue
        sp = round(md_val * MD_TO_SP_RATIO, 1)
        lines.append(f"  - {label}: {md_val} MD / {sp} SP")

    prompt = (
        "The following effort estimates are available for a software project:\n"
        + "\n".join(lines) + "\n\n"
        "Generate exactly 3 key insights comparing these estimates. Combine insights if necessary to stay at 3. "
        "Focus on gaps between methods, accuracy implications, and actionable recommendations.\n"
        "Return ONLY a JSON array — no markdown fences, no explanation:\n"
        '[{"type": "green|yellow|blue", "title": "<short title>", "description": "<1-2 sentences>"}]'
    )

    system = "You are a project estimation analyst. Return only valid JSON."
    raw = ai_respond(system, [], prompt).strip()
    
    if not raw:
        raise ValueError("AI returned an empty response instead of insights JSON.")

    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
    except json.JSONDecodeError:
        pass
        
    extracted = _extract_json_block(raw, is_array=True)
    if extracted is not None:
        return extracted
            
    raise ValueError(f"AI returned invalid JSON for insights. Response snippet: {raw[:200]}")


def _extract_metric_via_ai(text: str, source_name: str) -> dict | None:
    """Analyze supplementary data using the LLM to determine total MD/SP."""
    if not text or not text.strip(): 
        return None
    logger.info("Extracting %s metrics from supplementary payload...", source_name)
    prompt = (
        f"Analyze the following {source_name} data (extracted as up to 3 columns) and determine the TOTAL generic effort in man-days (MD).\n\n"
        f"Data snippet (first 4000 characters):\n{text[:4000]}\n\n"
        'Return EXACTLY one JSON object with no markdown fences, explanation, or extra keys. Use this format: {"md": <integer>, "sp": <float>}'
    )
    raw = ai_respond("You are an expert at extracting effort metrics into JSON.", [], prompt)
    try:
        res = _parse_json_response(raw)
        if isinstance(res, dict) and res.get("md") is not None:
            md_val = float(res["md"])
            sp_val = float(res.get("sp", round(md_val * MD_TO_SP_RATIO, 1)))
            return {"md": md_val, "sp": sp_val}
    except Exception as e:
        logger.warning("Failed to extract %s metrics: %s", source_name, e)
    return None

def _calculate_accuracy(md_val: float, actual_md: float) -> float:
    """Calculate accuracy percentage based on actual_md baseline."""
    return round(max(0, min(100, (1 - abs(md_val - actual_md) / actual_md) * 100)), 1)


# ── Routes ─────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", md_to_sp_ratio=MD_TO_SP_RATIO)


@app.route("/estimate", methods=["POST"])
def estimate():
    # ── 1. Collect file uploads ───────────────────────────────────────────
    brd_file          = request.files.get("brd")
    epic_list_file    = request.files.get("epic_list")
    jira_effort_file  = request.files.get("jira_effort")
    bitbucket_pr_file = request.files.get("bitbucket_pr")

    # ── 2. Collect text fields ────────────────────────────────────────────
    stakeholder_md_raw = request.form.get("stakeholder_md", "").strip()
    try:
        stakeholder_md_val = float(stakeholder_md_raw) if stakeholder_md_raw else None
    except ValueError:
        stakeholder_md_val = None

    # Restore Epic List mandatory check
    if not epic_list_file or not epic_list_file.filename:
        if brd_file and brd_file.filename:
            return jsonify({"error": "An Epic List is required to estimate. Please upload the Epic List (the BRD alone is not sufficient)."}), 400
        return jsonify({"error": "Please upload an Epic List to begin estimation."}), 400

    try:
        # ── 3. Extract Epic List text (optional for now) ──────────────────
        epic_list_text = ""
        if epic_list_file and epic_list_file.filename:
            epic_list_text = save_and_extract(epic_list_file, "epic")
            if not epic_list_text.strip():
                return jsonify({"error": "Could not extract any text from the Epic List document."}), 400

        # ── 4. Extract BRD as optional background context ─────────────────
        brd_text = save_and_extract(brd_file, "brd") if (brd_file and brd_file.filename) else ""

        # Must have at least one source
        if not epic_list_text.strip() and not brd_text.strip():
            return jsonify({"error": "Please upload a BRD or an Epic List to begin estimation."}), 400

        # ── 5. Extract other optional supplementary files ─────────────────
        jira_effort_text = save_and_extract(jira_effort_file, "jira")
        bitbucket_text   = save_and_extract(bitbucket_pr_file, "bitbucket")
        has_jira_effort  = bool(jira_effort_text)
        has_bitbucket    = bool(bitbucket_text)

        # ── 6. Chunk BRD (optional) and Epic List for AI processing ───────────
        brd_chunks  = chunk_text(brd_text, chunk_size=BRD_CHUNK_SIZE) if brd_text.strip() else []
        epic_chunks = chunk_text(epic_list_text) if epic_list_text.strip() else []
        if brd_chunks:
            logger.info("BRD: %d chars → %d chunks (size=%d)",
                        len(brd_text), len(brd_chunks), BRD_CHUNK_SIZE)
        if epic_chunks:
            logger.info("Epic List: %d chars → %d chunks", len(epic_list_text), len(epic_chunks))

        # ── 7. Load prompts ───────────────────────────────────────────────
        basic_prompt          = load_prompt("basic_prompt.txt")
        advanced_instructions = load_prompt("advanced_instructions.txt")


        # ── 8. Build finalize messages ────────────────────────────────────
        n = len(epic_chunks)
        if n > 0:
            FINALIZE_BASIC = (
                "You have now received the COMPLETE Epic List.\n"
                "Produce a basic effort estimate (no organizational context) for the TOTAL project.\n"
                "CRITICAL: You MUST return ONLY this JSON right now — no markdown fences, no explanation, no 'OK':\n"
                '{"project_name": "<inferred name>", "ai_basic": {"md": <integer>, "sp": <float>}}'
            )
        else:
            FINALIZE_BASIC = (
                "Produce a basic effort estimate (no organizational context) for the TOTAL project based entirely on the provided BRD context.\n"
                "CRITICAL: Do not reply with 'OK'. You MUST return ONLY this JSON right now — no markdown fences, no explanation:\n"
                '{"project_name": "<inferred name>", "ai_basic": {"md": <integer>, "sp": <float>}}'
            )

        # Contextual uses org rules only — actuals handled separately
        FINALIZE_CONTEXTUAL = (
            "You have already produced a basic estimate. Now apply ADVANCED organizational rules.\n\n"
            f"{advanced_instructions}\n\n"
            "CRITICAL INSTRUCTION: You MUST calculate a NEW total MD value. DO NOT simply copy your previous MD estimate. "
            "Add the % buffers and multipliers to your previous MD value, then recalculate SP.\n\n"
            "Return ONLY this JSON — no markdown fences, no explanation, no 'OK':\n"
            '{"project_name": "<inferred name>", "ai_contextual": {"md": <integer>, "sp": <float>}}'
        )

        # ── 9. Estimation: BRD summarisation + Epic List pass ─────────────────
        logger.info("Starting estimation…")
        basic_result, contextual_result = run_estimation(
            basic_prompt, brd_chunks, epic_chunks, FINALIZE_BASIC, FINALIZE_CONTEXTUAL
        )

        if not isinstance(basic_result, dict):
            logger.warning("basic_result was not a dict, defaulting to {}")
            basic_result = {}
        if not isinstance(contextual_result, dict):
            logger.warning("contextual_result was not a dict, defaulting to {}")
            contextual_result = {}

        # ── 8. Build available_data dict for insight generation ───────────
        project_name = (
            contextual_result.get("project_name")
            or basic_result.get("project_name", "")
        )

        ai_basic      = basic_result.get("ai_basic", {})
        if not isinstance(ai_basic, dict): ai_basic = {}
        
        # The model might stubbornly return "ai_basic" again during the contextual phase 
        # because of the history injection. Check both keys to be safe.
        ai_contextual = contextual_result.get("ai_contextual") or contextual_result.get("ai_basic", {})
        if not isinstance(ai_contextual, dict): ai_contextual = {}
        # ── 8. Extract Metrics via AI for supplementary files ───────────
        jira_scope = None
        jira_actual = _extract_metric_via_ai(jira_effort_text, "Jira Actual Effort") if has_jira_effort else None
        bitbucket = _extract_metric_via_ai(bitbucket_text, "Bitbucket PR Effort") if has_bitbucket else None

        # Collect whichever estimates we actually have for insight generation
        available_data = {}
        if stakeholder_md_val is not None:
            available_data["Stakeholder Estimate"] = stakeholder_md_val
        if ai_basic.get("md") is not None:
            available_data["AI Estimate (Basic)"] = ai_basic["md"]
        if ai_contextual.get("md") is not None:
            available_data["AI Estimate (Contextual)"] = ai_contextual["md"]
        if jira_scope and jira_scope.get("md") is not None:
            available_data["Jira Scope"] = jira_scope["md"]
        if jira_actual and jira_actual.get("md") is not None:
            available_data["Jira Actual Effort"] = jira_actual["md"]
        if bitbucket and bitbucket.get("md") is not None:
            available_data["Bitbucket Coding Effort"] = bitbucket["md"]

        # ── 9. Generate insights from available data ──────────────────────
        insights = []
        if len(available_data) >= 2:
            logger.info("Generating insights from: %s", list(available_data.keys()))
            try:
                insights = generate_insights(available_data)
            except Exception as exc:
                logger.warning("Insight generation failed: %s", exc)

        # ── 10. Compute accuracy (only if jira_actual known) ─────────────
        accuracy = {}
        actual_md = jira_actual.get("md") if jira_actual else None
        if actual_md:
            if stakeholder_md_val is not None:
                accuracy["stakeholder"] = _calculate_accuracy(stakeholder_md_val, actual_md)
            if ai_basic.get("md"):
                accuracy["ai-basic"] = _calculate_accuracy(ai_basic["md"], actual_md)
            if ai_contextual.get("md"):
                accuracy["ai-contextual"] = _calculate_accuracy(ai_contextual["md"], actual_md)
            if jira_scope and jira_scope.get("md"):
                accuracy["jira-scope"] = _calculate_accuracy(jira_scope["md"], actual_md)

        # ── 11. Build response (only include fields with real data) ───────
        response = {
            "project_name":  project_name,
            "ai_basic":      ai_basic,
            "ai_contextual": ai_contextual,
            "insights":      insights,
        }
        if accuracy:
            response["accuracy"] = accuracy
        if jira_scope:
            response["jira_scope"] = jira_scope
        if jira_actual:
            response["jira_actual"] = jira_actual
        if bitbucket:
            response["bitbucket"] = bitbucket

        return jsonify(response)

    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI returned invalid JSON: {e}"}), 500
    except Exception as e:
        logger.exception("Estimation failed")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000, threaded=True, use_reloader=False)

